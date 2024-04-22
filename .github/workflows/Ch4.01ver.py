import openpyxl
path = r"C:\Users\Admin\Desktop\Paloci Tracker\LogBook.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

Ref="A"+str(len(sheet["A"]))
AttemptN=(sheet[Ref].value+1)

import os
fileName = os.path.basename(__file__)

from datetime import datetime,date
today = date.today()
Date = today.strftime("%m/%d/%y")
now = datetime.now()
Time = now.strftime("%H:%M:%S")

sheet.append([AttemptN,fileName, Date, Time])
workbook.save(path)
workbook.close()


print()
print('------------------------Atempt Number %s ------------------------'%(AttemptN+1))
print()



###prevents hard codeing the password value###
PasswordPath=r'C:\Users\Admin\Desktop\Paloci Tracker\AhabAppPassword.txt'
AhabPasswordFile=open(PasswordPath,"r")
AhabPassword = AhabPasswordFile.read()
AhabPasswordFile.close()
###----------------------------------------###

###establishes email protocall##
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
def harpoon(subject,message):
    # Email configuration
    me = "captianahab.tips@gmail.com"
    my_password = AhabPassword
    you = 'captianahab.tips@gmail.com'
    Realyou = "14.chris.oh@gmail.com"

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = me
    msg['To'] = you

    html = '<html><body><p>%s</p></body></html>'%message
    part2 = MIMEText(html, 'html')

    msg.attach(part2)

    # Send the message via gmail's regular server, over SSL - passwords are being sent, afterall
    s = smtplib.SMTP_SSL('smtp.gmail.com')
    # uncomment if interested in the actual smtp conversation
    # s.set_debuglevel(1)
    # do the smtp auth; sends ehlo if it hasn't been sent already
    s.login(me, my_password)

    s.sendmail(me, you, msg.as_string())
    s.quit()
        #Example:Harpoon('*thunk*','Yes, making this a function does work')
###---------------------------###


###allows for a current time to be printed###
def CurrentTime():
    today = date.today()
    RelativeDate = today.strftime("%m/%d/%y")
    now = datetime.now()
    RelativeTime = now.strftime("%H:%M:%S")
    return RelativeTime
###---------------------------------------###


###Current Date for the subject to be printed###
def RelativeDateFunct():
    today = date.today()
    RelativeDate = today.strftime("%m/%d/%y")
    return RelativeDate
###---------------------------------------###

year= today.strftime("20%y")



from inspect import currentframe, getframeinfo
#(getframeinfo(currentframe()).lineno, "code")


###Determimine the year of the previous FD report to prevent a jan 1 problem###
LineIDReport=(getframeinfo(currentframe()).lineno)
IncumbantFilePath=r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 2 Website To PDF\Overview\Temp\Incumbant\IncumbantFD.xlsx"
workbook = openpyxl.load_workbook(IncumbantFilePath)
sheet = workbook.active
IncumbantYear=(sheet['H3'].value)
workbook.save(IncumbantFilePath)
workbook.close()

if IncumbantYear == None: 
    harpoon(RelativeDateFunct(),'Program Stopped:Incumbant File Blank \nOof, looks like somethign went wrong,\nUse an old backup of the incumbant file and run it again.\n\nCurrent Time: %s:\nThe line refrence ID is %s \n\n\n*burp*'%(CurrentTime(),LineIDReport))
        #email notif
    print('none test triggered')
    exit()
        #ends program
    

if not(IncumbantYear == year):
    harpoon(RelativeDateFunct(),'Program Stop: YEARi<>YEARc\nThere is a problem with the year recognised was "%s", compared to the year identified %s.\nCurrent Time: %s.\nThe line refrence ID is %s.\n\n\n Sorry Cheif'%(IncumbantYear,year,CurrentTime(),LineIDReport))
    print('program stop due to incumbant year break')
    print('Program Stop: YEARi<>YEARc\nThere is a problem with the year recognised was "%s", compared to the year identified %s.\nCurrent Time: %s.\nThe line refrence ID is %s.\n\n\n Sorry Cheif'%(IncumbantYear,year,CurrentTime(),LineIDReport))
    exit()
    #ends program
###-------------------------------------------------------------------------###

import  time, os,requests
import xml.etree.ElementTree as ET
IncumbantLength=0
ChalangerLength=0
    
ChalangerFDSet=[]
## WHILE THE INCUMBANT AND THE CHALANGER FILE ARE THE SAME LENGTH STAY IN A HOLDING PATTERN###
while IncumbantLength ==  ChalangerLength:
    
    year= today.strftime("20%y")
        # this needs to be repeated in case this program runs into a new year, it will have a problem with downloading prevous year's reports 
        
    ###Given a URL, this will download the FD annual report as a Zip File###
    url = str('https://disclosures-clerk.house.gov/public_disc/financial-pdfs/%sFD.zip'%(year))
            #comes up with the url for the year's fincial disclosure summary

    LineIDReport=(getframeinfo(currentframe()).lineno, "code")
    
    Ch2ZipOutPut = r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 2 Website To PDF\Overview\Temp"
        #the url of the new PTR
    os.makedirs(Ch2ZipOutPut, exist_ok=True)
        # Create the directory if it doesn't exist
    filename = os.path.join(Ch2ZipOutPut, url.split("/")[-1])
        # Get the filename from the URL

    response = requests.get(url)
        # Send a GET request to the URL

    if response.status_code == 200:
        # Check if the request was successful
        # Write the PDF content to a file
        with open(filename, 'wb') as f:
            f.write(response.content)
    #       print(f"Zip Summary downloaded successfully to {filename}")
    else:
        print("Failed to download the Zip Summary")
        harpoon(RelativeDateFunct(),'Program Stop: Zip Download Error\n:\n Something is wrong with the process to download the zipfile from the congressional website.\nCurrent Time: %s.\nThe line refrence ID is %s.\n\n\n Sorry Cheif'%(CurrentTime(),LineIDReport))
        #email notif
        exit()
        #ends program
    ###------------------------------------------------------------------###



    ###Taking the zip file just downloaded, and extracting an XLM file from it###
    from zipfile import ZipFile
    XML_File_path=''
    XML_Name=''
    with ZipFile(filename, 'r') as zip:
        #open the zip file in 'r' Read mode
    #    zip.printdir()
        #print all contents of the zip file
        XML_File_path=r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 2 Website To PDF\Overview\Temp\XML Chalanger Folder"
        XML_Name='%sFD.xml'%year
        zip.extract(member=XML_Name,path=XML_File_path)
            #puts the xml file we just created into the xml holder
    os.remove(filename)
            #deleatesthe zip file we just made

    ChalangerXMLPath='%s\Chalanger-%s'%(XML_File_path,XML_Name)
    os.rename('%s\%s'%(XML_File_path,XML_Name),ChalangerXMLPath)
            # identifies the XML file we just made as the "chalanger file"
    ###-----------------------------------------------------------------------###


    ###Taking the Element Tree from the XML File and Converting it to a list of lists, with an output for the lenght###
    tree = ET.parse(ChalangerXMLPath)
    root = tree.getroot()
    j=-1
    l=0
    FDDetail=[]
    ChalangerFDSet=[]
    for i in root:
        j=j+1
        FDDetail=[root[j][8].text]
        for c in range(8):
            FDDetail.append(root[j][c].text)
        ChalangerFDSet.append(FDDetail)
    #print('DocID - Prefix - LName - FName - Suffix - FilingType - StateDst - Year - FilingDate')
    #print(*ChalangerFDSet,sep="\n")
    ChalangerLength=len(ChalangerFDSet)
    print('Chalanger Length is:'+str(ChalangerLength))
    os.remove(ChalangerXMLPath)
    ###-------------------------------------------------------------------------------------------------------------###



    ### Pulling the lentth and year of the Incumbant FD ###
    IncumbantFilePath=r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 2 Website To PDF\Overview\Temp\Incumbant\IncumbantFD.xlsx"
    workbook = openpyxl.load_workbook(IncumbantFilePath)
    sheet = workbook.active
    IncumbantLength = (len(sheet["A"])-2)
    IncumbantYear=(sheet['H3'].value)
    workbook.save(IncumbantFilePath)
    workbook.close()
    print (IncumbantLength)
    #print (IncumbantYear)
    ###----------------------------------------###

    ###Wait and check for a developement in 3 hours###
    if IncumbantLength ==  ChalangerLength:
        
        print('no changes as of %s adding 1 to sim a change\n, will sleep for 3 hrs'%Time)
#       time.sleep(10800)    
        harpoon(RelativeDateFunct(),'All Quiet on the western front.\n(turned off) Program will rest for 3 hrs\nBoth Fincial Disclosure(FD) reports have a length of %s.\n Chalanger:%s and Incumbant:%s\nCurrent Time: %s:\nThe line refrence ID is %s \n\n\n'%(IncumbantLength,ChalangerLength,IncumbantLength,CurrentTime(),LineIDReport))
            #email notif
    ###--------------------------------------------###
###-----------------------------------------------------------------------------------------###


###Random check to cover the 3 possible inequalities###
if IncumbantLength > ChalangerLength:
    harpoon(RelativeDateFunct(),'Program Stopped:Len(i)>Len(c) \nThe incumbant FD report has more items than the Chalanger FD... How?\nNot sure how this is possible if there the years agree.\nNext steps wise I have nothing for you...\n\nCurrent Time: %s:\nThe line refrence ID is %s \n\n\nSorry'%(CurrentTime(),LineIDReport))
        #email notif
    print('Incumbant Length: %s, Chalanger Length: %s'%(IncumbantLength,ChalangerLength))
    print('Program Stopped:Len(i)>Len(c) \nThe incumbant FD report has more items than the Chalanger FD... How?\nNot sure how this is possible if there the years agree.\nNext steps wise I have nothing for you...\n\nCurrent Time: %s:\nThe line refrence ID is %s \n\n\nSorry'%(CurrentTime(),LineIDReport))
    exit()
        #ends program
###-------------------------------------------------###
    

###Here is where we assemble the two lists to make the compairiion easier, this first list is stored in an excel file so that will make things... less easy###
IncumbantFilePath=r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 2 Website To PDF\Overview\Temp\Incumbant\IncumbantFD.xlsx"
workbook = openpyxl.load_workbook(IncumbantFilePath)
sheet = workbook.active
IncumbantLength = (len(sheet["A"])-2)
IncumbantYear=(sheet['H3'].value)
    
IncumbantIDSet=[]
for i in range(IncumbantLength):
    ref='A'+str((2+i))
    IncumbantIDSet.append(sheet[ref].value)
workbook.save(IncumbantFilePath)
workbook.close()   
###--------------------------------------------------------------------------------------------------------------------------------------------------------###


###takes list of trades and spits out the paired FD IDs and their Trade Type abrevi###
Chalanger_Pairred_Set=[]
for i in range(len(ChalangerFDSet)):
    holding=[]
    holding.append(ChalangerFDSet[i][0])
    holding.append(ChalangerFDSet[i][5])
    Chalanger_Pairred_Set.append(holding)

ChalangerIDSet=[]
for i in range(len(ChalangerFDSet)):
    ChalangerIDSet.append(ChalangerFDSet[i][0])

Discrepancy = list(set(ChalangerIDSet).difference(IncumbantIDSet))

OutGoingMail=[]
for i in Discrepancy:
    OutGoingMail.append(Chalanger_Pairred_Set[int(ChalangerIDSet.index(i))])
print(OutGoingMail)
###--------------------------------------------------------------------------------###


###Deleating the old Incumbant Excel File, Creating a new one, adds data### 
IncumbantFilePath=r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 2 Website To PDF\Overview\Temp\Incumbant\IncumbantFD.xlsx"
os.remove(IncumbantFilePath)
    #deletes old Incumbant file
from openpyxl import  Workbook 
BlankWorkBook = Workbook()
BlankWorkBook.save(IncumbantFilePath)
BlankWorkBook.close()
#print('waiting 30 s to see if the incumbant file deleted, it could be causing the fail') #nope, it's not the deleting, that works
for i in range(1):
    print(i)
    time.sleep(1)
    #creates new Excel File

workbook = openpyxl.load_workbook(IncumbantFilePath)
sheet = workbook.active
sheet.append(['DocID','Prefix','LName','FName','Suffix','FilingType','StateDst','Year','FilingDate'])
for i in range(len(ChalangerFDSet)):
    sheet.append(ChalangerFDSet[i])
workbook.save(IncumbantFilePath)
workbook.close()
                #adds new data
###---------------------------------------------------------------###

PDFPathPackage=[]
 
for i in range(len(OutGoingMail)):
    subdirectory = 'financial-pdfs'
    if OutGoingMail[i][1] == 'P':
        subdirectory = 'ptr-pdfs'
    FilingID=OutGoingMail[i][0]
    
    ###Download the file from a given link, and store it as a pdf###
    url = str('https://disclosures-clerk.house.gov/public_disc/%s/%s/%s.pdf'%(subdirectory,year,FilingID))
                    #What about 
                    #      https://disclosures-clerk.house.gov/public_disc/financial-pdfs/2009/8140068.pdf
                                                                        #^^^^^^^^^ does not conform with
    Ch2PdfOutPut = r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 2 Website To PDF\Output"
        #the url of the new PTR
    os.makedirs(Ch2PdfOutPut, exist_ok=True)
        # Create the directory if it doesn't exist
    filename = os.path.join(Ch2PdfOutPut, url.split("/")[-1])
        # Get the filename from the URL
    response = requests.get(url)
        # Send a GET request to the URL
    if response.status_code == 200:
        # Check if the request was successful
        # Write the PDF content to a file
        with open(filename, 'wb') as f:
            f.write(response.content)
            print(f"PDF downloaded successfully to {filename}")
            print('¯\_(ツ)_/¯')
    else:
        print("Failed to download the PDF")
        harpoon(RelativeDateFunct(),'Program Stopped: Failed To download one of the PDFs \nThe Line From Out Going Mail is %s\n\nCurrent Time: %s:\nThe line refrence ID is %s \n\n\n[klind little blurb]'%(OutGoingMail[i],CurrentTime(),LineIDReport))
            #email notif
        exit()
        #ends program
        #LineIDReport=(getframeinfo(currentframe()).lineno, "code")
    
    
    PdfPath=("C:\\Users\\Admin\\Desktop\\Paloci Tracker\\Ch 2 Website To PDF\\Output\\%s.pdf"%(FilingID))
    PDFPathPackage.append(PdfPath)
    
    ###----------------------------------------------------------###

print(PDFPathPackage)

###--------------------------#taken from 1.6--------------------------###
from pypdf import PdfReader
import re

for i in range(len(PDFPathPackage)):
    GivenReportPath=PDFPathPackage[i]

    PdfReceptacle=""
    reader = PdfReader(GivenReportPath,"rb")
    PdfLength=len(reader.pages)
    for p in range(PdfLength):
        page = reader.pages[p]

        for c in page.extract_text():
            PdfReceptacle=PdfReceptacle+(c)
                # adds new text char to the lsit, also allowing for the new paragraphs to become new paragraphs and not [\n]s

    
    PdfReceptacle =re.sub("P\W\W\W\W\W\W\WT\W\W\W\W\W\W\W\W\W\WR\W\W\W\W\W","Periodic Transaction Report",PdfReceptacle)
    PdfReceptacle =re.sub("F\W\W\W\WI\W\W\W\W\W\W\W\W\W\W","\nFiller Information\n",PdfReceptacle)
    PdfReceptacle =re.sub("Status:","\nStatus:",PdfReceptacle)
    PdfReceptacle =re.sub("T\W\W\W\W\W\W\W\W\W\W\W","\nTransactions\n",PdfReceptacle)
    PdfReceptacle =re.sub("F\W\W\W\W\WS\W\W\W\W\W","\nFiling Status",PdfReceptacle)
    PdfReceptacle =re.sub("S\W\W\W\W\W\W\W\W\WO\W","Subholding Of",PdfReceptacle)
    PdfReceptacle =re.sub("A\W\W\W\WC\W\W\W\WD\W\W\W\W\W\W","\nAsset Class Details",PdfReceptacle)
    PdfReceptacle =re.sub("L\W\W\W\W\W\W\W","Location",PdfReceptacle)
    PdfReceptacle =re.sub("Filing ID","\nFiling ID",PdfReceptacle)
    PdfReceptacle =re.sub("I\W\W\W\W\W\WP\W\W\W\W\WO\W\W\W\W\W\W\W\W","\nInital Public Offerings ",PdfReceptacle)
    PdfReceptacle =re.sub("C\W\W\W\W\W\W\W\W\W\W\W\W \W\W\WS\W\W\W\W\W\W\W\W","\nCertification and Signature\n",PdfReceptacle)
    PdfReceptacle =re.sub(",","",PdfReceptacle)
        #changes the unprintable charecters and some formatting

    #print (PdfReceptacle)

    FilingID=""
    FName =""
    LName = ""
    Status = ""
    StateAndDistrict=""
    AssetTradeName=""
    Ticker=""
    AssetType=""
    TransactionType=""
    TransactionDate=""
    NotificaionDate=""
    AmmountRangeLow=""
    AmmountRangeUpper=""


    DetailIndex = int(PdfReceptacle.index('Filing ID #'))
    FilingID = int(PdfReceptacle[DetailIndex+11:DetailIndex+19])


    DetailIndex = int(PdfReceptacle.index('Name:'))
    DetailIndex2 = int(PdfReceptacle.index('Status:'))
    NameList = PdfReceptacle[DetailIndex+5:DetailIndex2].split()
    FName = NameList[1]
    LName = NameList[-1]

    DetailIndex = int(PdfReceptacle.index('Status:'))
    DetailIndex2 = int(PdfReceptacle.index('State/District:'))
    Status = PdfReceptacle[DetailIndex+7:DetailIndex2]

    DetailIndex = int(PdfReceptacle.index('State/District:'))
    DetailIndex2 = int(PdfReceptacle.index('Transactions'))
    StateAndDistrict = PdfReceptacle[DetailIndex+15:DetailIndex2-1]

    DetailIndex = int(PdfReceptacle.index('$200?'))
    DetailIndex2 = int(PdfReceptacle.index('(',DetailIndex))
    AssetTradeName=(PdfReceptacle[DetailIndex+9:DetailIndex2-1])

    DetailIndex = DetailIndex2
    DetailIndex2 = int(PdfReceptacle.index(')',DetailIndex))
    Ticker=(PdfReceptacle[DetailIndex+1:DetailIndex2])

    DetailIndex = DetailIndex2
    DetailIndex2 = int(PdfReceptacle.index(']',DetailIndex))
    AssetType=(PdfReceptacle[DetailIndex+3:DetailIndex2])
                                                                                                                                #Hmmm these can prob be made into a function...later problem though

    TransactionType=(PdfReceptacle[DetailIndex2+2])

    TransactionDate=(PdfReceptacle[DetailIndex2+4:DetailIndex2+14])

    NotificaionDate=(PdfReceptacle[DetailIndex2+16:DetailIndex2+25])

    DetailIndex = int(PdfReceptacle.index('$',DetailIndex2))
    DetailIndex2 = int(PdfReceptacle.index('-',DetailIndex))
    AmmountRangeLow=int(PdfReceptacle[DetailIndex+1:DetailIndex2-1])


    DetailIndex = int(PdfReceptacle.index('$',DetailIndex2))
    DetailIndex2 = int(PdfReceptacle.index('Filing',DetailIndex))
    AmmountRangeUpper=int(PdfReceptacle[DetailIndex+1:DetailIndex2-1])


    path = r"C:\Users\Admin\Desktop\Paloci Tracker\Ch 1  Pdf To Excel\Output\OutputValveCh1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    DollarSignCount=PdfReceptacle.count('$')
    NTrades=int((((DollarSignCount))/2))

    sheet.append([AttemptN,FilingID, FName, LName, Status,StateAndDistrict,AssetTradeName, Ticker, AssetType, TransactionType, TransactionDate, NotificaionDate, AmmountRangeLow, AmmountRangeUpper,1,NTrades])

    DollarSignCount=PdfReceptacle.count('$')
    NTrades=int((((DollarSignCount))/2))
    for i in range(NTrades-1):
            #howto det t or not? 
                    # when detail index 2 reaches end of page 
        print(i+1)
        DetailIndex = int(PdfReceptacle.index('SP',DetailIndex2))
        DetailIndex2 = int(PdfReceptacle.index('(',DetailIndex))
        AssetTradeName=(PdfReceptacle[DetailIndex+3:DetailIndex2-1])

        DetailIndex = DetailIndex2
        DetailIndex2 = int(PdfReceptacle.index(')',DetailIndex))
        Ticker=(PdfReceptacle[DetailIndex+1:DetailIndex2])
        
        DetailIndex = DetailIndex2
        DetailIndex2 = int(PdfReceptacle.index(']',DetailIndex))
        AssetType=(PdfReceptacle[DetailIndex+3:DetailIndex2])

        DetailIndex = int(PdfReceptacle.index('/',DetailIndex2))
        TransactionType=(PdfReceptacle[DetailIndex-4])

        DetailIndex = int(PdfReceptacle.index('/',DetailIndex2))
        DetailIndex2 = int(PdfReceptacle.index('/',DetailIndex+1))
        TransactionDate=(PdfReceptacle[DetailIndex2-5:DetailIndex2+5])

        DetailIndex = int(PdfReceptacle.index('/',DetailIndex2+1))
        DetailIndex2 = int(PdfReceptacle.index('/',DetailIndex+1))
        NotificaionDate=(PdfReceptacle[DetailIndex2-5:DetailIndex2+5])

        DetailIndex = int(PdfReceptacle.index('$',DetailIndex2))
        DetailIndex2 = int(PdfReceptacle.index('-',DetailIndex))
        AmmountRangeLow=int(PdfReceptacle[DetailIndex+1:DetailIndex2-1])
        
        DetailIndex = int(PdfReceptacle.index('$',DetailIndex2))
        DetailIndex2 = int(PdfReceptacle.index('Filing',DetailIndex))
        AmmountRangeUpper=int(PdfReceptacle[DetailIndex+1:DetailIndex2-1])
        
        SeriesNumber = i+2
        sheet.append([AttemptN,FilingID, FName, LName, Status,StateAndDistrict,AssetTradeName, Ticker, AssetType, TransactionType, TransactionDate, NotificaionDate, AmmountRangeLow, AmmountRangeUpper, SeriesNumber, NTrades])
        workbook.save(path)
        workbook.close()



#print('\n\n..............\nFilingID= %s.\n Name= %s. %s.\n status= %s.\n StateAndDistrict= %s.\n AssetTradeName= %s. \n Ticker= %s.\n AssetType= %s.\n TransactionType=%s \n TransactionDate= %s.\nNotificaionDate= %s.\nAmmountRangeLow= %s.\nAmmountRangeUpper= %s.'%(FilingID, FName, LName, Status,StateAndDistrict,AssetTradeName, Ticker, AssetType, TransactionType, TransactionDate, NotificaionDate, AmmountRangeLow, AmmountRangeUpper))


###-------------------------------------------------------------------###



#       harpoon(RelativeDateFunct(),'[title] \n[details]\n[extra details]\n[sugested next steps]\nCurrent Time: %s:\nThe line refrence ID is %s \n\n\n[klind little blurb]'%(CurrentTime(),LineIDReport))
        #email notif
#        exit()
        #ends program
        #LineIDReport=(getframeinfo(currentframe()).lineno, "code")
  

